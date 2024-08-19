import os
import sqlite3
# import openpyxl
import openpyxl
from tqdm import tqdm

from mconst import *

from mzweb import *

base_name = 'mus2024spb.db'


def export_to_sqlite():
    '''Экспорт данных из xlsx в sqlite'''
    '''2402 Экспорт данных из xlsx в sqlite'''

    # 1. Создание и подключение к базе

    # Получаем текущую папку проекта
    prj_dir = os.path.abspath(os.path.curdir)
    a = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    # Имя базы

    # метод sqlite3.connect автоматически создаст базу, если ее нет
    connect = sqlite3.connect(prj_dir + '/' + base_name)
    # курсор - это специальный объект, который делает запросы и получает результаты запросов
    cursor = connect.cursor()

    # создание таблицы если ее не существует
    cursor.execute(
        'CREATE TABLE IF NOT EXISTS spb91 (WKT text, mName text,googlelink text,description  text,style text,access text,citywallslink text,years text, architect text)')

    # 2. Работа c xlsx файлом

    # Читаем файл и лист1 книги excel
    file_to_read = openpyxl.load_workbook('spb91.xlsx', data_only=True)
    sheet = file_to_read['Стили_и_Entrances']

    # Цикл по строкам начиная со второй (в первой заголовки)

    for row in range(2, sheet.max_row + 1):
        # Объявление списка
        data = []
        # Цикл по столбцам от 1 до 4 ( 5 не включая)
        for col in range(1, 10):
            # value содержит значение ячейки с координатами row col
            value = sheet.cell(row, col).value
            # Список который мы потом будем добавлять
            data.append(value)

        # 3. Запись в базу и закрытие соединения

        # Вставка данных в поля таблицы
        cursor.execute("INSERT INTO spb91 VALUES (?, ?, ?, ?,?, ?, ?, ?,?);",
                       (data[0], data[1], data[2], data[3], data[4], data[5], data[6], data[7], data[8]))

    # сохраняем изменения
    connect.commit()
    # закрытие соединения
    connect.close()


def clear_base():
    '''Очистка базы sqlite'''
    '''2402 Очистка базы sqlite'''

    # Получаем текущую папку проекта
    prj_dir = os.path.abspath(os.path.curdir)

    # Имя базы
    base_name = 'auto.sqlite3'

    connect = sqlite3.connect(prj_dir + '/' + base_name)
    cursor = connect.cursor()

    # Запись в базу, сохранение и закрытие соединения
    cursor.execute("DELETE FROM spb91")
    connect.commit()
    connect.close()


def spbsql2table():
    '''2403 Берем из sql link, получаем данные через web и обновляем в sql'''

    class spbb:
        def __init__(self, d1, d2, d3, d4, d5, d6, d7, d8, d9):
            self.d1 = d1
            self.d2 = d2
            self.d3 = d3
            self.d4 = d4
            self.d5 = d5
            self.d6 = d6
            self.d7 = d7
            self.d8 = d8
            self.d9 = d9

        #

    prj_dir = os.path.abspath(os.path.curdir)
    a = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

    # метод sqlite3.connect автоматически создаст базу, если ее нет
    connection = sqlite3.connect(prj_dir + '/' + base_name)
    # курсор - это специальный объект, который делает запросы и получает результаты запросов
    cursor = connection.cursor()

    cursor.execute('SELECT * FROM spb91')

    web = cWebm()

    ar = []

    for row in cursor.fetchall():
        ar.append(spbb(row[0], row[1], row[2], row[3], row[4], row[5], row[6], row[7], row[8]))

    import re
    mdesc = row[3]

    for ii in ar:
        ss = str(ii.d7)
        if ss.find("citywalls") < 0:
            mdesc = ii.d4

            match = re.search('https://www.citywalls.ru(.+?)html', mdesc)
            # if match:
            #     print(match.group(0))  #  "Keywords: key, key, key"
            #
            if match == None:
                print(ii.d2 + " " + "CHECK DESC")
                continue

            mcity = match.group(0)

            march, myaears, mstyle = web.getcitywalls(mcity)

            cursor.execute(
                'UPDATE spb91 SET years = ?, style =?, architect =?,citywallslink=?'
                'WHERE mName= ?',
                (myaears, mstyle, march, mcity, ii.d2,))
            connection.commit()
            res = "DONE"
        else:
            res = "SKIP"
        print(ii.d2 + " " + res)


# Запуск функции
# export_to_sqlite()

# web = cWebm()

def spbsql2jpg():
    '''2406 Проверка. Несколько строк из базы в jpg файлы'''

    prj_dir = os.path.abspath(os.path.curdir)
    a = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    connection = sqlite3.connect(prj_dir + '/' + base_name)
    cursor = connection.cursor()

    ll = cursor.execute('SELECT mainphoto from  buildingdb')

    for i in range(3):
        ablob = cursor.fetchone()[0]
        filename = "spbsql2jpg"
        with open(filename + str(i) + ".jpg", 'wb') as output_file:
            output_file.write(ablob)


def spblink2sql(mlist, web, cursor):
    '''2406 на входе link, на выходе - строка msql'''

    progress_bar = tqdm(mlist, desc='Processing cities', total=len(mlist))

    for i in range(len(mlist)):
        md = web.getcitywalls2gethouse(mlist[i])
        cursor.execute(
            'INSERT INTO buildingdb (city,name,year,style,status,'
            'arch1,arch2,arch3,arch4,'
            'addr1n,addr2n,addr3n,addr4n,'
            'link,lastdate,lasttime,'
            'longitude,latitude,note)'
            ' VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)',
            (md["city"], md["name"], md["year"], md["style"], md["status"],
             md["arch1"], md["arch2"], md["arch3"], md["arch4"],
             md["maddr1"], md["maddr2"], md["maddr3"], md["maddr4"],
             mlist[i], getcdate(), getctime(),
             md["longitude"], md["latitude"], md["note"]))
        cursor.connection.commit()

        ind = cursor.lastrowid
        # print("IIIIIIII "+str(ind))

        cursor.execute(
            'INSERT INTO photodb (building_fk,'
            'comment,'
            'photo,photo_link,'
            'photo_small,photo_small_link,'
            'is_main)'
            ' VALUES (?,?,?,?,?,?,?)',
            (ind,
             md["photocomment"],
             md["photo"], md["photo_link"],
             md["photo_small"], md["photo_small_link"],
             1))
        cursor.connection.commit()

        progress_bar.update()


def check_isHouseInDB(mlist, cursor):
    msql = 'SELECT link FROM buildingdb'

    cursor.execute(msql)

    mlinks = []

    for row in cursor.fetchall():
        mlinks.append(row[0])

    mres = []

    for i in mlist:
        if i not in mlinks:
            mres.append(i)
        # else:
        #     print("INF: DUPLICATE link" + i)

    print('Начальное значение:%d из них новые:%d' % (len(mlist), len(mres)))

    return mres


def getcitywalls2getstreetv2(cursor, web, link):
    tlist = web.getcitywalls2getstreet(link)
    mres = check_isHouseInDB(tlist, cursor)

    return mres


def main():
    mlist = []

    web = cWebm()

    prj_dir = os.path.abspath(os.path.curdir)
    a = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    connection = sqlite3.connect(prj_dir + '/' + base_name)
    cursor = connection.cursor()

    # # 240724 DONEALL Мойка мосты
    # mlist = mlist + getcitywalls2getstreetv2(cursor, web, "https://www.citywalls.ru/search-street1777.html")
    # 240724 DONEALL Мойка набережная
    # mlist = mlist + getcitywalls2getstreetv2(cursor, web, "https://www.citywalls.ru/search-street99.html")
    # # 240724 DONEALL  7 линия ВО
    # mlist = mlist + getcitywalls2getstreetv2(cursor, web, "https://www.citywalls.ru/search-street37.html")
    # # 240724 DONEALL  8 линия ВО
    # mlist = mlist + getcitywalls2getstreetv2(cursor, web, "https://www.citywalls.ru/search-street23.html")
    # # 240724 DONEALL  09 линия ВО
    # mlist = mlist + getcitywalls2getstreetv2(cursor, web, "https://www.citywalls.ru/search-street36.html")
    # # 240724 DONEALL  10 линия ВО
    # mlist = mlist +getcitywalls2getstreetv2(cursor, web, "https://www.citywalls.ru/search-street7.html")
    # # 240724 DONEALL  11 линия ВО
    # mlist = mlist + getcitywalls2getstreetv2(cursor, web, "https://www.citywalls.ru/search-street8.html")
    # # 240724 DONEALL   12 линия ВО
    # mlist = mlist + getcitywalls2getstreetv2(cursor, web, "https://www.citywalls.ru/search-street38.html")
    # # 240724 DONEALL   13 линия ВО
    # mlist = mlist + getcitywalls2getstreetv2(cursor, web, "https://www.citywalls.ru/search-street53.html")
    # #  240724 DONEALL   Чайковского
    # mlist = mlist + getcitywalls2getstreetv2(cursor, web, "https://www.citywalls.ru/search-street1.html")

    # mlist.append("https://www.citywalls.ru/house32411.html")
    # mlist.append("https://www.citywalls.ru/house772.html")


    # # # 240724 DONEALL Дворцовая набережная
    # mlist = mlist + getcitywalls2getstreetv2(cursor, web, "https://www.citywalls.ru/search-street69.html")
    #
    # # # 240724 DONEALL Кутузова набережная
    # mlist = mlist + getcitywalls2getstreetv2(cursor, web, "https://www.citywalls.ru/search-street210.html")

    # # 240720DONE Воскресенская набережная
    #mlist = mlist + getcitywalls2getstreetv2(cursor, web, "https://www.citywalls.ru/search-street3455.html")


    # # 240724 DONEALL Большой проспект ВО
    # mlist = mlist + getcitywalls2getstreetv2(cursor, web, "https://www.citywalls.ru/search-street32.html")
    #
    # # 240724 DONEALL Средний  проспект ВО
    # mlist = mlist + getcitywalls2getstreetv2(cursor, web, "https://www.citywalls.ru/search-street14.html")
    #
    # # 240724 DONEALL Малый  проспект ВО
    # mlist = mlist + getcitywalls2getstreetv2(cursor, web, "https://www.citywalls.ru/search-street35.html")
    #
    # # 240724 DONEALL  Репина
    # mlist = mlist + getcitywalls2getstreetv2(cursor, web, "https://www.citywalls.ru/search-street159.html")
    # 240811 DONEALL Английская
    # mlist = mlist + getcitywalls2getstreetv2(cursor, web, "https://www.citywalls.ru/search-street208.html")
    # # 240811 DONEALL Шмидта
    # mlist = mlist + getcitywalls2getstreetv2(cursor, web, "https://www.citywalls.ru/search-street55.html")
    # # 240811 DONEALL  Университетская
    # mlist = mlist + getcitywalls2getstreetv2(cursor, web, "https://www.citywalls.ru/search-street39.html")
    # #  240811 DONEALL Галерная
    # mlist = mlist + getcitywalls2getstreetv2(cursor, web, "https://www.citywalls.ru/search-street217.html")
    # #  240811 DONEALL 14
    # mlist = mlist + getcitywalls2getstreetv2(cursor, web, "https://www.citywalls.ru/search-street34.html")
    # #  240811 DONEALL 15
    # mlist = mlist + getcitywalls2getstreetv2(cursor, web, "https://www.citywalls.ru/search-street54.html")
    # #  240811 DONEALL 16
    # mlist = mlist + getcitywalls2getstreetv2(cursor, web, "https://www.citywalls.ru/search-street27.html")
    # #  240811 DONEALL 17
    # mlist = mlist + getcitywalls2getstreetv2(cursor, web, "https://www.citywalls.ru/search-street56.html")
    # #  240811 DONEALL 18
    # mlist = mlist + getcitywalls2getstreetv2(cursor, web, "https://www.citywalls.ru/search-street33.html")
    # #  240811 DONEALL 19
    # mlist = mlist + getcitywalls2getstreetv2(cursor, web, "https://www.citywalls.ru/search-street57.html")
    # #  240811 DONEALL 20
    # mlist = mlist + getcitywalls2getstreetv2(cursor, web, "https://www.citywalls.ru/search-street47.html")
    # #  240811 DONEALL 21
    # mlist = mlist + getcitywalls2getstreetv2(cursor, web, "https://www.citywalls.ru/search-street43.html")
    # #  240811 DONEALL 22
    # mlist = mlist + getcitywalls2getstreetv2(cursor, web, "https://www.citywalls.ru/search-street46.html")

    # #  240811 DONEALL Щербаков
    # mlist = mlist + getcitywalls2getstreetv2(cursor, web, "https://www.citywalls.ru/search-street587.html")

    # #  Большой ПС
    #   mlist = mlist + getcitywalls2getstreetv2(cursor, web, "https://www.citywalls.ru/search-street97.html")

    #  Канонера
    mlist = mlist + getcitywalls2getstreetv2(cursor, web, "https://www.citywalls.ru/search-street799.html")



    # #  240811 DONEALL Графский
    # mlist = mlist + getcitywalls2getstreetv2(cursor, web, "https://www.citywalls.ru/search-street491.html")
    #  240811 DONEALL Владимирский
    # mlist = mlist + getcitywalls2getstreetv2(cursor, web, "https://www.citywalls.ru/search-street170.html")
    # #  240811 DONEALL Загородный
    # mlist = mlist + getcitywalls2getstreetv2(cursor, web, "https://www.citywalls.ru/search-street491.html")
    # #  240811 DONEALL Рубинщт
    # mlist = mlist + getcitywalls2getstreetv2(cursor, web, "https://www.citywalls.ru/search-street88.html")
    # #  240811 DONEALL Ломо
    # mlist = mlist + getcitywalls2getstreetv2(cursor, web, "https://www.citywalls.ru/search-street478.html")
    # # 240811 DONEALL Джамб
    # mlist = mlist + getcitywalls2getstreetv2(cursor, web, "https://www.citywalls.ru/search-street399.html")
    # #  240811 DONEALL Бородинска
    # mlist = mlist + getcitywalls2getstreetv2(cursor, web, "https://www.citywalls.ru/search-street324.html")
    # #  240811 DONEALL бол Каз
    #  mlist = mlist + getcitywalls2getstreetv2(cursor, web, "https://www.citywalls.ru/search-street478.html")
    # #  240811 DONEALL Звниг
    # mlist = mlist + getcitywalls2getstreetv2(cursor, web, "https://www.citywalls.ru/search-street177.html")
    # #  240811 DONEALL Правд
    # mlist = mlist + getcitywalls2getstreetv2(cursor, web, "https://www.citywalls.ru/search-street84.html")
    # #  240811 DONEALL Соци
    # mlist = mlist + getcitywalls2getstreetv2(cursor, web, "https://www.citywalls.ru/search-street31.html")
    # #  240811 DONEALL мара
    # mlist = mlist + getcitywalls2getstreetv2(cursor, web, "https://www.citywalls.ru/search-street75.html")
    # #  240811 DONEALL Разь
    # mlist = mlist + getcitywalls2getstreetv2(cursor, web, "https://www.citywalls.ru/search-street76.html")

    # !!!!!!!!!!!!!!!!!!!!#  мал Каз - 1 СТАРНИЦА
    # mlist = mlist + getcitywalls2getstreetv2(cursor, web, "https://www.citywalls.ru/search-street478.html")

    # !!!!!!!!!!!!!!!!!!!!#  Горо - ГЛЮК
    # mlist = mlist + getcitywalls2getstreetv2(cursor, web, "https://www.citywalls.ru/search-street94.html")


    # # ------------------------------------------------
    # # 23
    # mlist = mlist + getcitywalls2getstreetv2(cursor, web, "https://www.citywalls.ru/search-street58.html")
    # # 24
    # mlist = mlist + getcitywalls2getstreetv2(cursor, web, "https://www.citywalls.ru/search-street106.html")
    # # 25
    # mlist = mlist + getcitywalls2getstreetv2(cursor, web, "https://www.citywalls.ru/search-street685.html")
    # # 26
    # mlist = mlist + getcitywalls2getstreetv2(cursor, web, "https://www.citywalls.ru/search-street688.html")
    # # 27
    # mlist = mlist + getcitywalls2getstreetv2(cursor, web, "https://www.citywalls.ru/search-street687.html")
    # # ------------------------------------------------

    # # 240724 DONEALL
    # mlist = mlist + getcitywalls2getstreetv2(cursor, web, "https://www.citywalls.ru/search-street13.html")

    # # 240724 DONEALL
    # mlist = mlist + getcitywalls2getstreetv2(cursor, web, "https://www.citywalls.ru/search-street2.html")


    # Садовая
    # # Error в https://www.citywalls.ru/house32411.html
    # mlist = mlist + getcitywalls2getstreetv2(cursor, web, "https://www.citywalls.ru/search-street101.html")

    # # Михайловская короткая - ПРОВЕРИТ ПОЧЕМУ ГЛЮК
    # mlist = mlist + getcitywalls2getstreetv2(cursor, web, "https://www.citywalls.ru/search-street237.html")
    #
    # # Энгельса
    # mlist = mlist + getcitywalls2getstreetv2(cursor, web, "https://www.citywalls.ru/search-street635.html")


    mres = check_isHouseInDB(mlist, cursor)

    print(mres)

    f = open('mreport__.txt', 'w', encoding="utf-8")
    for line in mres:
        f.write(f"{line}\n")
    f.close()

    spblink2sql(mres, web, cursor)


# spbsql2table()

if __name__ == "__main__":
    main()
